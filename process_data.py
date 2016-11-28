import openpyxl
from process_data_functions import generate_modified_date_header_names, generate_modified_date_column_ints, generate_header_names, generate_exported_files_list, generate_modified_date_header_names
from openpyxl import Workbook #Interact with Excel
import os #Handle file names and paths

class Generate_Column_Data(object):
	"""
	Input a single/list of excel column(s) and return ints and strs.
	"""
	def __init__(self, columns_as_a_list, column_with_id=""):
		# Convert to a list if not.
		if not (isinstance(columns_as_a_list, list)):
			columns_as_a_list = [columns_as_a_list]

		# If the list is strings, generate .columns_str and convert to .columns_int
		if (isinstance(columns_as_a_list[0], str)):
			self.columns_str = columns_as_a_list
			modified_list_of_columns = []
			for column in columns_as_a_list:
				modified_column = openpyxl.utils.column_index_from_string(column)
				modified_list_of_columns.append(modified_column)
			self.columns_int = modified_list_of_columns

		# If the list is ints, generate .columns_int and convert to .columns_str
		if (isinstance(columns_as_a_list[0], int)):
			# If the list is strings, generate .columns_str and convert to .columns_int
			self.columns_int = columns_as_a_list

			modified_list_of_columns = []
			for column in columns_as_a_list:
				modified_column = openpyxl.utils.get_column_letter(column)
				modified_list_of_columns.append(modified_column)
			self.columns_str = modified_list_of_columns

		#If there is a user id column entered
		if (column_with_id != ""):
			if (isinstance(columns_as_a_list[0], str)):
				#Convert id Column to int
				self.column_with_id_str = column_with_id
				self.column_with_id_int = openpyxl.utils.column_index_from_string(column_with_id)

			if (isinstance(columns_as_a_list[0], int)):
				# Convert id Column to str
				self.column_with_id_int = column_with_id
				self.column_with_id_str = openpyxl.utils.get_column_letter(column_with_id)

			try:
				columns_int = self.columns_int[:]
				columns_str = self.columns_str[:]
				columns_int.remove(self.column_with_id_int)
				columns_str.remove(self.column_with_id_str)
				self.columns_int_no_id = columns_int
				self.columns_str_no_id = columns_str
			except(ValueError):
				print("ID column not found in list")
				return None
		else:
			self.columns_int_no_id = self.columns_int
			self.columns_str_no_id = self.columns_str

		#Generate [0,1,2...
		self.indecies = []
		for i in range(0, len(self.columns_int)): self.indecies.append(i)

		#Generate [0,1,2... minus the ID column
		self.indecies_no_id = []
		for i in range(0, len(self.columns_int)): self.indecies_no_id.append(i)
		if(column_with_id != ""):
			self.indecies_no_id.remove(columns_as_a_list.index(column_with_id))
		else:
			self.indecies_no_id = self.indecies
		print("yes")

##################################################################
#User Defined Settings############################################
##################################################################
directory_with_exported_files = "Exports"
filename_of_merged_excel_docs = "MergedResults.xlsx"
workbook_sheet_name_of_memrged_excel_doc = "Sheet1"
columns_to_merge_together = [1,3,4] #As list or, single string, single int
column_with_unique_id = 1 #as single string, single int, no lists

##################################################################
#Setup Base File, Merged File, and Directory of Exported Files####
##################################################################
exported_files_directory = directory_with_exported_files
exported_files_list = generate_exported_files_list(exported_files_directory)

#Base is the starting file for matching all newer files. It is the oldest file in the folder. 
base_file = exported_files_directory + "/" + exported_files_list[0]
base_file_name = os.path.basename(base_file)[:-5]
base_workbook = openpyxl.load_workbook(filename = base_file)
base_sheet = base_workbook['Sheet1']

#Merged will be where all the information is combined together
merged_file = filename_of_merged_excel_docs
merged_workbook = openpyxl.Workbook()
merged_sheet = merged_workbook.active 
merged_sheet.title = workbook_sheet_name_of_memrged_excel_doc

##################################################################
#Define Columns###################################################
##################################################################
#Type in list of columns to use which will also generate a list of the names of the headers
base = Generate_Column_Data(columns_to_merge_together, column_with_id = column_with_unique_id)
base_headers = generate_header_names(base_sheet,base.columns_str)

#Generate a list for the merged columns
#1 to base.columns_int is space for the copied base columns. base.columns_int to base.columns_int_no_id is
# for columns for date modified columns.
merged_columns = []
for i in range(1,len(base.columns_int + base.columns_int_no_id)+1): merged_columns.append(i)
merged = Generate_Column_Data(merged_columns)

merged_headers = generate_modified_date_header_names(base_headers,base.indecies_no_id)
print(merged_headers)


#merged_date_created_column = generate_string_from_column_index(max(base_ints)*2 + 1) #add this column after the last column of the modified file  (len(base) + len(modified_columns)) gives the last column, +1 to get an empty column
#merged_date_created_header = "Created Date"


#Copy the values of the headers in the base_sheet to the merged_sheet

##################################################################
#Add Columns to Merge Doc#########################################
##################################################################

#Add base headers to merged_sheet
for column in base.columns_str:
	index = 0
	merged_sheet[column + "1"].value = base_headers[index]
	index +=1




merged_workbook.save(merged_file)

"""
#Add date modified headers to merged_sheet
for index in range(0,len(merged_date_modified_columns_ints)):
	merged_sheet[merged_date_modified_columns[index]+ "1"].value = merged_date_modified_headers[index]


merged_sheet[merged_date_created_column + "1"] = merged_date_created_header

#Copy the values of the cells in the base_sheet to the merged_sheet
for index_row in range(2,base_sheet.max_row+1):
	for index_col in base:
		merged_sheet[index_col+str(index_row)].value = base_sheet[index_col+str(index_row)].value
	merged_sheet[merged_date_created_column + str(index_row)] = base_file_name
"""
"""

def loop_through_non_base_files(exported_files_list): 
	#loop through all files not used to generate base file by using 1:
	for exported_file in exported_files_list[1:]:
		print(exported_files_list)
		#Loop will be where all the information for the current Excel doc will be stored
		loop_file = exported_files_directory + "/" + exported_file
		loop_file_name = os.path.basename(loop_file)[:-5]
		loop_workbook = openpyxl.load_workbook(filename = loop_file)
		loop_sheet = loop_workbook['Sheet1']

		#For each row in the loop file, compare it against the Merged file by using
		#The ID as a comparison. If a user ID from the loop file is found in the merged
		#file, comparisons of answers will be made. Otherwise Continue on.
		#print("Comparing " + loop_file + " and " + base_file)
		loop_rows_not_found = []
		for loop_row in range(2,loop_sheet.max_row+1):
			id_row_found = False #Once the id_row match is found, send this variable to true
			for merged_row in range(2,merged_sheet.max_row+1):
				#print("Comparing loop:" + str(loop_sheet[id_column + str(loop_row)].value) + " and merged:" + str(merged_sheet[id_column + str(merged_row)].value))
				if(loop_sheet[id_column + str(loop_row)].value == merged_sheet[id_column + str(merged_row)].value):
					#print(merged_sheet[id_column + str(merged_row)].value)
					id_row_found = True

					#In this space will be a function that compares each 

					for base_column in base_without_id:
						base_column_int = generate_column_index_from_string(base_column)
						if(loop_sheet[base_column + str(loop_row)].value != merged_sheet[base_column + str(merged_row)].value ):
							#print(base_column_int)
							#print(len_of_merged_columns)
							merged_sheet[generate_string_from_column_index(base_column_int[0] + len_of_merged_columns) + str(merged_row)].value = loop_file_name
							merged_sheet[base_column + str(merged_row)].value = loop_sheet[base_column + str(loop_row)].value
					break
			if(id_row_found != True):
				loop_rows_not_found.append(loop_row)
		#print(loop_rows_not_found)

		# For the rows that were not found, append them to the end of the document. 
		merged_sheet_empty_last_row = merged_sheet.max_row+1
		for loop_row_not_found in loop_rows_not_found:
			for  index_col in base:
				 merged_sheet[index_col+str(merged_sheet_empty_last_row)].value = loop_sheet[index_col+str(loop_row_not_found)].value
			merged_sheet[merged_date_created_column + str(merged_sheet_empty_last_row)].value = loop_file_name
			merged_sheet_empty_last_row += 1
		merged_workbook.save(merged_file)	

loop_through_non_base_files(exported_files_list)

"""

#Check whether the above cells match the expected string contents
"""
error_count = 0
for cell_key, cell_value in header_validation.items():
	if(base_sheet[cell_key].value != header_validation[cell_key]):
		print("There was an error in sheet " + base_file + " at " + cell_key + 
			  "\n'" + base_sheet[cell_key].value + "' does not match '" + header_validation[cell_key] +"'\n")
		error_count +=1
if(error_count != 0):
	print("Fix " + str(error_count) +" error(s) before continuing.")
"""


